import httpx
import os
import pandas as pd
from pandas.tseries.offsets import BDay
from datetime import datetime
from send_price_bbg.conf import fund_names_dict

from prefect import task, flow, get_run_logger
from openpyxl import load_workbook

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from prefect.blocks.system import Secret
from prefect.variables import Variable

SMTP_HOST = "smtp.sendgrid.net"
SMTP_USER = "apikey"
API_KEY = Secret.load("sendgrid-api-key").get()
FROM_EMAIL = "ops@3commafunds.com"
TO_EMAIL = "dduarte@3commafunds.com"

RECIPIENTS = Variable.get("price_recipients")
RECIPIENTS = ["nhaga5@gmail.com", "dduarte@3commafunds.com"]

def map_data_bloomberg(priceList, names_dict):
    priceList = priceList.reset_index()

    priceList = priceList[["priceDate", "identifier", "price", "volume"]]
    priceList["fundName"] = priceList["identifier"].map(names_dict)
    priceList["currency"] = "EUR"
    priceList["classAssets"] = round(priceList["price"] * priceList["volume"], 4)
    priceList["fundSize"] = round(
        priceList.groupby("fundName")["classAssets"].transform("sum"), 4
    )
    priceList["bid"] = ""
    priceList["offer"] = ""
    final_prices = priceList[
        [
            "priceDate",
            "identifier",
            "fundName",
            "currency",
            "price",
            "bid",
            "offer",
            "fundSize",
            "classAssets",
            "volume",
        ]
    ]

    return final_prices


@task
def get_prices():
    url = "https://3comma-helpers.up.railway.app/fund/history"
    response = httpx.get(url)
    if response.status_code != 200:
        raise Exception(f"Failed to fetch data: {response.status_code}")

    json_data = response.json()
    return pd.DataFrame(json_data).set_index("identifier", drop=True)


@task
def get_day_up_prices(identifier, day=datetime.today() - BDay(1)):
    today_str = day.date().strftime("%Y-%m-%dT%H:%M:%S")

    prices = get_prices()

    if identifier != "all":
        prices = prices.loc[identifier]

    return prices[prices["priceDate"] == today_str]


@task
def create_excel(identifier, day=datetime.today() - BDay(1)):
    price_list = get_day_up_prices(identifier, day)

    priceList = map_data_bloomberg(price_list, fund_names_dict)
    priceList = priceList.astype(str)

    template_excel_file = "send_price_bbg/3CC_Template.xlsx"
    sheet_name = "Price Format"
    book = load_workbook(template_excel_file)

    if sheet_name not in book.sheetnames:
        book.create_sheet(sheet_name)

    sheet = book[sheet_name]
    start_row = 9
    start_col = 1

    for row_idx, row in enumerate(priceList.values, start=start_row):
        for col_idx, value in enumerate(row, start=start_col):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    return book


@task
def gen_simple_email(_from, recipients, subject, body, attachment=None):
    print(";".join(recipients))
    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"] = _from
    #msg['To'] = 'ops@3commafunds.com'
    #msg["BCC"] = ", ".join(recipients)
    msg['To'] = ";".join(recipients)
    msg['BCC'] = "nhaga5@gmail.com"

    msg.attach(MIMEText(body, "plain"))

    if attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(open(attachment, "rb").read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{attachment}"')
        msg.attach(part)

    return msg


@flow(name="05 Send Fund Prices", log_prints=True)
def send_fund_prices_email(signature="3 Comma Capital Funds", day_lag=1):
    subject = "Daily Unit Pricing for Global Crypto Fund"
    body = f"""
    Hi all,

    The prices for 3 Comma Capital Funds are hereby attached.

    Thanks,
    {signature}
    """

    filename = "3CC_Fund_Prices.xlsx"
    create_excel("all").save(filename)

    msg = gen_simple_email(
        _from="ops@3commafunds.com",
        recipients=RECIPIENTS, #["dduarte@3commafunds.com"],
        subject=subject,
        body=body,
        attachment=filename,
    )
    logger = get_run_logger()

    try:
        with smtplib.SMTP_SSL(SMTP_HOST, 465) as server:
            server.login(SMTP_USER, API_KEY)
            server.sendmail(FROM_EMAIL, TO_EMAIL, msg.as_string())
        logger.info("Email sent successfully")
    except Exception as e:
        logger.error(f"Failed to send email: {e}")

    os.remove(filename)


if __name__ == "__main__":
    send_fund_prices_email()

    # secret_block = Secret.load("gmail-app-pass").get()
    # print(secret_block)

import httpx
import pandas as pd
from prefect import flow, task
from prefect.blocks.system import Secret
from datetime import datetime
from pandas.tseries.offsets import BDay
from openpyxl import load_workbook
import yagmail
import os
from conf import fund_names_dict

secret_block = Secret.load("sender-email-password")


def map_data_bloomberg(priceList, names_dict):
    priceList = priceList.reset_index()

    priceList = priceList[["priceDate", "identifier", "price", "volume"]]
    priceList["fundName"] = priceList["identifier"].map(names_dict)
    priceList["currency"] = "EUR"
    priceList["classAssets"] = round(priceList["price"] * priceList["volume"], 4)
    priceList["fundSize"] = round(
        priceList.groupby("fundName")["classAssets"].transform("sum"), 4
    )
    priceList["bid"] = ""
    priceList["offer"] = ""
    final_prices = priceList[
        [
            "priceDate",
            "identifier",
            "fundName",
            "currency",
            "price",
            "bid",
            "offer",
            "fundSize",
            "classAssets",
            "volume",
        ]
    ]

    return final_prices


@task
def get_prices():
    url = "https://3comma-helpers.up.railway.app/fund/history"
    response = httpx.get(url)
    if response.status_code != 200:
        raise Exception(f"Failed to fetch data: {response.status_code}")

    json_data = response.json()
    return pd.DataFrame(json_data).set_index("identifier", drop=True)


@task
def get_day_up_prices(identifier, day=datetime.today() - BDay(1)):
    today_str = day.date().strftime("%Y-%m-%dT%H:%M:%S")

    prices = get_prices()

    if identifier != "all":
        prices = prices.loc[identifier]

    return prices[prices["priceDate"] == today_str]


@task
def create_excel(identifier, day=datetime.today() - BDay(1)):
    price_list = get_day_up_prices(identifier, day)

    priceList = map_data_bloomberg(price_list, fund_names_dict)
    priceList = priceList.astype(str)

    template_excel_file = "3comma_flows/3CC_Template.xlsx"
    new_file = "3CC_Fund_Prices.xlsx"
    sheet_name = "Price Format"
    book = load_workbook(template_excel_file)

    if sheet_name not in book.sheetnames:
        book.create_sheet(sheet_name)

    sheet = book[sheet_name]
    start_row = 9
    start_col = 1

    for row_idx, row in enumerate(priceList.values, start=start_row):
        for col_idx, value in enumerate(row, start=start_col):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    return book

    book.save(new_file)


@flow(log_prints=True)
def send_email(receiver_email, signature, filename="3CC_Fund_Prices.xlsx"):
    sender_email = os.environ.get("EMAIL_SENDER")
    alias_dict = {f"{sender_email}": "ops@3commafunds.com"}
    password = secret_block.get()

    yag = yagmail.SMTP(alias_dict, password)

    subject = "Daily Unit Pricing for Global Crypto Fund"
    body = f"""
    Hi all,

    The prices for 3 Comma Capital Funds are hereby attached.

    Thanks,
    {signature}
    """

    create_excel("all").save(filename)

    try:
        yag.send(
            bcc=receiver_email,
            subject=subject,
            contents=body,
            attachments=filename,
        )
        print("Email sent successfully with attachments!")
    except Exception as e:
        print(f"Failed to send email: {e}")


if __name__ == "__main__":
    # recipients = [
    #     "ops@3commafunds.com",
    #     "fundpricing@bloomberg.net",
    #     "nav@morningstareurope.com",
    #     "lipper.pricing@lseg.com"
    # ]
    recipients = "dduarte@3commafunds.com"
    signature = "David Duarte"

    send_email(recipients, signature)
